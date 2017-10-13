from PIL import Image
import openpyxl as exel
from random import shuffle
from openpyxl.styles import PatternFill



im = Image.open("test.jpg") #change to desired file name
width, height = im.size

new_height = 140        #change to desired pixel height
new_width  = int((new_height * width )/ height)

img = im.resize((new_width, new_height), Image.ANTIALIAS)

pix = img.load()
print(width,height)
rgblist = [0,1,2]
wb = exel.load_workbook("b1.xlsx")
sheet = wb.get_sheet_names()
sheet1 = wb.get_sheet_by_name(sheet[0])

for  j in range(new_width-1):
    for i in range(0,(new_height-1)):
        for k in range(3):
            a = sheet1.cell(row=k+3*i+1, column=j+1, value  =pix[j,i][k])

            str1 = str(hex(pix[j, i][rgblist[k]]))[2:]
            if len(str1)<2:
                str1="0"+str1
            col = "FF"+("00"*k)+str1+"00"*(2-k)

            a.fill = PatternFill(bgColor = col,fill_type = "solid")
img.save("test1.jpg")
im.close()
wb.save("test1.xlsx") #change to desired output name
wb.close()
print(0)